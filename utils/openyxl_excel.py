from openpyxl import load_workbook
import time
from utils.log import get_path


class parseExcel(object):
    def __init__(self, excelPath):
        self.excelPath = excelPath
        self.workbook = load_workbook(excelPath)  # 加载excel
        self.sheet = self.workbook['Sheet1']  # 获取第一个sheet

    # 设置当前要操作的sheet对象，使用index来获取相应的sheet
    def set_sheet_by_index(self, sheet_index):
        sheet_name = self.workbook.get_sheet_names()[sheet_index]
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)
        return self.sheet

    # 获取当前默认sheet的名字
    def get_default_sheet(self):
        return self.sheet.title

    # 设置当前要操作的sheet对象，使用sheet名称来获取相应的sheet
    def set_sheet_by_name(self, sheet_name):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        self.sheet = sheet
        return self.sheet

    # 获取默认sheet中最大的行数
    def get_max_row_no(self):
        return self.sheet.max_row

    # 获取默认 sheet 的最大列数
    def get_max_col_no(self):
        return self.sheet.max_column

    # 获取默认sheet的最小（起始）行号
    def get_min_row_no(self):
        return self.sheet.min_row

    # 获取默认sheet的最小（起始）列号
    def get_min_col_no(self):
        return self.sheet.min_column

    # 获取默认 sheet 的所有行对象，
    def get_all_rows(self):
        return list(self.sheet.iter_rows())
        # return list(self.rows)也可以

    # 获取默认sheet中的所有列对象
    def get_all_cols(self):
        return list(self.sheet.iter_cols())
        # return list(self.sheet.columns)也可以

    # 从默认sheet中获取某一列，第一列从0开始
    def get_single_col(self, col_no):
        return self.get_all_cols()[col_no]

    # 从默认sheet中获取某一行，第一行从0开始
    def get_single_row(self, row_no):
        return self.get_all_rows()[row_no]

    # 从默认sheet中，通过行号和列号获取指定的单元格，注意行号和列号从1开始
    def get_cell(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no)

    # 从默认sheet中，通过行号和列号获取指定的单元格中的内容，注意行号和列号从1开始
    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入指定内容，注意行号和列号从1开始
    # 调用此方法的时候，excel不要处于打开状态
    def write_cell_content(self, row_no, col_no, content, font=None):
        self.sheet.cell(row=row_no, column=col_no).value = content
        self.workbook.save(self.excelPath)
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入当前日期，注意行号和列号从1开始
    # 调用此方法的时候，excel不要处于打开状态
    def write_cell_current_time(self, row_no, col_no):
        time1 = time.strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.cell(row=row_no, column=col_no).value = str(time1)
        self.workbook.save(self.excelPath)
        return self.sheet.cell(row=row_no, column=col_no).value

    def save_excel_file(self):
        self.workbook.save(self.excelPath)

    def test(self):
        return '230281SM3M1L'


if __name__ == '__main__':
    p = parseExcel(get_path() + '/企业导入模版.xlsx')
    p.write_cell_content(2, 9, '1234567')
    # print("获取默认行：",p.get_default_sheet())
    # print("取得行号和列号(2,4)单元格：",p.get_cell_content(2,4))
    # card_pass = p.test()
    # rows = p.get_all_rows()
    # list_value = []
    # for cell in rows:
    #     for res in cell:
    #         list_value.append(res.value)
    # if card_pass in list_value:
    #     print('删除')
    # else:
    #     print('激活')
